"""
Excel文档解析器

解析Excel文档内容，提取工作表、表格数据、表单字段等信息
"""

from pathlib import Path
from typing import Dict, List, Any, Optional
from src.utils.db_manager import node_state
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

from ..core import BaseParser, DocumentType, ProcessingStatus
from ..config import get_settings
from ..utils.logger import get_logger, performance_monitor, error_handler


class ExcelParser(BaseParser):
    """Excel文档解析器"""
    
    def __init__(self):
        super().__init__()
        self.logger = get_logger(__name__)
        self.settings = get_settings()
        
        if not HAS_OPENPYXL:
            self.logger.warning("openpyxl库未安装，无法处理.xlsx格式文件")
        
        if not HAS_XLRD:
            self.logger.warning("xlrd库未安装，无法处理.xls格式文件")
        
        if not HAS_WIN32COM:
            # self.logger.warning("win32com库未安装，某些功能可能不可用")
            pass
    
    def get_supported_formats(self) -> List[DocumentType]:
        """返回支持的文档格式列表"""
        return [DocumentType.XLS, DocumentType.XLSX]
    
    def validate_input(self, input_data: Any) -> bool:
        """验证输入数据的有效性"""
        if isinstance(input_data, (str, Path)):
            file_path = Path(input_data)
            return (file_path.exists() and 
                    file_path.suffix.lower() in ['.xls', '.xlsx'])
        return False
    
    def process(self, input_data: Any) -> Any:
        """处理输入数据并返回结果"""
        if isinstance(input_data, (str, Path)):
            return self.parse(Path(input_data))
        else:
            raise ValueError(f"不支持的输入类型: {type(input_data)}")
    
    @error_handler()
    @performance_monitor()
    def parse(self, file_path: Path) -> Dict[str, Any]:
        """
        解析Excel文档
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            解析结果字典
        """
        self.logger.info(f"{node_state}-=-填表员===开始解析Excel文档: {file_path}")
        
        try:
            result = {
                'file_path': str(file_path),
                'file_size': file_path.stat().st_size,
                'document_type': self._get_excel_type(file_path).value,
                'metadata': {},
                'worksheets': [],
                'text_content': '',
                'tables': [],
                'forms': [],
                'charts': [],
                'named_ranges': [],
            }
            
            # 根据文件扩展名选择解析方法
            if file_path.suffix.lower() == '.xlsx':
                self._parse_xlsx(file_path, result)
            else:
                self._parse_xls(file_path, result)
            
            # 检测表单字段
            self._detect_form_fields(result)
            
            self.status = ProcessingStatus.SUCCESS
            self.logger.info(f"{node_state}-=-填表员===Excel文档解析完成: {file_path}")
            
            return result
            
        except Exception as e:
            self.status = ProcessingStatus.FAILED
            self.logger.error(f"Excel文档解析失败 {file_path}: {e}")
            raise
    
    def _parse_xlsx(self, file_path: Path, result: Dict[str, Any]):
        """解析XLSX格式文档"""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl库未安装，无法解析XLSX文件")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 提取元数据
            result['metadata'] = self._extract_xlsx_metadata(workbook)
            
            # 解析工作表
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_info = self._parse_xlsx_worksheet(worksheet, sheet_name)
                result['worksheets'].append(sheet_info)
                
                # 累积文本内容
                sheet_text = self._extract_sheet_text(sheet_info)
                result['text_content'] += f"工作表 {sheet_name}:\n{sheet_text}\n\n"
            
            # 提取命名范围
            result['named_ranges'] = self._extract_xlsx_named_ranges(workbook)
            
            # 关闭工作簿
            workbook.close()
            
        except Exception as e:
            self.logger.error(f"解析XLSX文件失败: {e}")
            raise
    
    def _parse_xls(self, file_path: Path, result: Dict[str, Any]):
        """解析XLS格式文档"""
        if not HAS_XLRD:
            self.logger.warning("xlrd库未安装，尝试使用COM接口解析XLS文件")
            self._parse_xls_com(file_path, result)
            return
        
        try:
            workbook = xlrd.open_workbook(str(file_path))
            
            # 提取元数据
            result['metadata'] = self._extract_xls_metadata(workbook)
            
            # 解析工作表
            for sheet_idx in range(workbook.nsheets):
                worksheet = workbook.sheet_by_index(sheet_idx)
                sheet_info = self._parse_xls_worksheet(worksheet, sheet_idx)
                result['worksheets'].append(sheet_info)
                
                # 累积文本内容
                sheet_text = self._extract_sheet_text(sheet_info)
                result['text_content'] += f"工作表 {worksheet.name}:\n{sheet_text}\n\n"
            
        except Exception as e:
            self.logger.error(f"解析XLS文件失败: {e}")
            # 尝试使用COM接口
            self._parse_xls_com(file_path, result)
    
    def _parse_xls_com(self, file_path: Path, result: Dict[str, Any]):
        """使用COM接口解析XLS文件"""
        if not HAS_WIN32COM:
            self.logger.error("win32com库未安装，无法解析XLS文件")
            return
        
        try:
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            try:
                workbook = excel_app.Workbooks.Open(str(file_path.absolute()))
                
                # 解析工作表
                for sheet in workbook.Worksheets:
                    sheet_info = self._parse_xls_worksheet_com(sheet)
                    result['worksheets'].append(sheet_info)
                    
                    # 累积文本内容
                    sheet_text = self._extract_sheet_text(sheet_info)
                    result['text_content'] += f"工作表 {sheet.Name}:\n{sheet_text}\n\n"
                
                workbook.Close()
                
            finally:
                excel_app.Quit()
                
        except Exception as e:
            self.logger.error(f"使用COM接口解析XLS文件失败: {e}")
    
    def _extract_xlsx_metadata(self, workbook: Workbook) -> Dict[str, Any]:
        """提取XLSX文档元数据"""
        metadata = {}
        
        try:
            props = workbook.properties
            metadata = {
                'title': props.title or '',
                'creator': props.creator or '',
                'description': props.description or '',
                'subject': props.subject or '',
                'keywords': props.keywords or '',
                'category': props.category or '',
                'created': str(props.created) if props.created else '',
                'modified': str(props.modified) if props.modified else '',
                'last_modified_by': props.lastModifiedBy or '',
                'revision': str(props.revision) if props.revision else '',
            }
            
            # 清理空值
            metadata = {k: v for k, v in metadata.items() if v}
            
        except Exception as e:
            self.logger.warning(f"提取XLSX元数据失败: {e}")
        
        return metadata
    
    def _extract_xls_metadata(self, workbook) -> Dict[str, Any]:
        """提取XLS文档元数据"""
        metadata = {}
        
        try:
            # xlrd提供的元数据有限
            metadata = {
                'sheet_count': workbook.nsheets,
                'sheet_names': workbook.sheet_names(),
            }
            
        except Exception as e:
            self.logger.warning(f"提取XLS元数据失败: {e}")
        
        return metadata
    
    def _parse_xlsx_worksheet(self, worksheet: Worksheet, sheet_name: str) -> Dict[str, Any]:
        """解析XLSX工作表"""
        sheet_info = {
            'name': sheet_name,
            'max_row': worksheet.max_row,
            'max_column': worksheet.max_column,
            'data': [],
            'merged_cells': [],
            'tables': [],
            'charts': [],
        }
        
        try:
            # 提取数据
            for row in worksheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                sheet_info['data'].append(row_data)
            
            # 提取合并单元格信息
            for merged_range in worksheet.merged_cells.ranges:
                sheet_info['merged_cells'].append(str(merged_range))
            
            # 检测表格结构
            tables = self._detect_tables_in_sheet(sheet_info['data'])
            sheet_info['tables'] = tables
            
        except Exception as e:
            self.logger.warning(f"解析工作表 {sheet_name} 失败: {e}")
        
        return sheet_info
    
    def _parse_xls_worksheet(self, worksheet, sheet_idx: int) -> Dict[str, Any]:
        """解析XLS工作表"""
        sheet_info = {
            'name': worksheet.name,
            'index': sheet_idx,
            'max_row': worksheet.nrows,
            'max_column': worksheet.ncols,
            'data': [],
            'tables': [],
        }
        
        try:
            # 提取数据
            for row_idx in range(worksheet.nrows):
                row_data = []
                for col_idx in range(worksheet.ncols):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell_value = str(cell.value) if cell.value is not None else ''
                    row_data.append(cell_value)
                sheet_info['data'].append(row_data)
            
            # 检测表格结构
            tables = self._detect_tables_in_sheet(sheet_info['data'])
            sheet_info['tables'] = tables
            
        except Exception as e:
            self.logger.warning(f"解析XLS工作表 {worksheet.name} 失败: {e}")
        
        return sheet_info
    
    def _parse_xls_worksheet_com(self, worksheet) -> Dict[str, Any]:
        """使用COM接口解析XLS工作表"""
        sheet_info = {
            'name': worksheet.Name,
            'data': [],
            'tables': [],
        }
        
        try:
            # 获取使用范围
            used_range = worksheet.UsedRange
            if used_range:
                # 提取数据
                values = used_range.Value
                if values:
                    if isinstance(values[0], tuple):
                        # 多行数据
                        for row in values:
                            row_data = [str(cell) if cell is not None else '' for cell in row]
                            sheet_info['data'].append(row_data)
                    else:
                        # 单行数据
                        row_data = [str(cell) if cell is not None else '' for cell in values]
                        sheet_info['data'].append(row_data)
            
            # 检测表格结构
            tables = self._detect_tables_in_sheet(sheet_info['data'])
            sheet_info['tables'] = tables
            
        except Exception as e:
            self.logger.warning(f"使用COM解析工作表 {worksheet.Name} 失败: {e}")
        
        return sheet_info
    
    def _extract_xlsx_named_ranges(self, workbook: Workbook) -> List[Dict[str, Any]]:
        """提取XLSX命名范围"""
        named_ranges = []
        
        try:
            for name in workbook.defined_names.definedName:
                named_range = {
                    'name': name.name,
                    'refers_to': str(name.attr_text),
                    'scope': name.localSheetId if name.localSheetId else 'workbook',
                }
                named_ranges.append(named_range)
                
        except Exception as e:
            self.logger.warning(f"提取命名范围失败: {e}")
        
        return named_ranges
    
    def _detect_tables_in_sheet(self, data: List[List[str]]) -> List[Dict[str, Any]]:
        """检测工作表中的表格结构"""
        tables = []
        
        if not data or len(data) < 2:
            return tables
        
        try:
            # 简单的表格检测算法
            # 查找连续的非空行作为表格
            table_start = None
            current_table_rows = []
            
            for row_idx, row in enumerate(data):
                # 检查行是否为空
                is_empty_row = all(cell.strip() == '' for cell in row)
                
                if not is_empty_row:
                    if table_start is None:
                        table_start = row_idx
                    current_table_rows.append(row)
                else:
                    # 空行，结束当前表格
                    if table_start is not None and len(current_table_rows) >= 2:
                        table_info = {
                            'start_row': table_start,
                            'end_row': table_start + len(current_table_rows) - 1,
                            'rows': len(current_table_rows),
                            'columns': len(current_table_rows[0]),
                            'data': current_table_rows,
                            'headers': current_table_rows[0] if current_table_rows else [],
                        }
                        tables.append(table_info)
                    
                    # 重置
                    table_start = None
                    current_table_rows = []
            
            # 处理文件末尾的表格
            if table_start is not None and len(current_table_rows) >= 2:
                table_info = {
                    'start_row': table_start,
                    'end_row': table_start + len(current_table_rows) - 1,
                    'rows': len(current_table_rows),
                    'columns': len(current_table_rows[0]),
                    'data': current_table_rows,
                    'headers': current_table_rows[0] if current_table_rows else [],
                }
                tables.append(table_info)
                
        except Exception as e:
            self.logger.warning(f"检测表格结构失败: {e}")
        
        return tables
    
    def _extract_sheet_text(self, sheet_info: Dict[str, Any]) -> str:
        """从工作表信息中提取文本内容"""
        text_lines = []
        
        try:
            for row in sheet_info.get('data', []):
                # 过滤空单元格，连接非空内容
                non_empty_cells = [cell for cell in row if cell.strip()]
                if non_empty_cells:
                    text_lines.append(' | '.join(non_empty_cells))
            
        except Exception as e:
            self.logger.warning(f"提取工作表文本失败: {e}")
        
        return '\n'.join(text_lines)
    
    def _detect_form_fields(self, result: Dict[str, Any]):
        """检测Excel文档中的表单字段"""
        form_fields = []
        
        try:
            # 在所有工作表中查找表单模式
            for sheet in result['worksheets']:
                sheet_text = self._extract_sheet_text(sheet)
                patterns = self._detect_form_patterns(sheet_text)
                
                for pattern in patterns:
                    form_field = {
                        'type': pattern['type'],
                        'label': pattern['matched_text'],
                        'sheet': sheet['name'],
                        'position': pattern['position'],
                        'required': False,
                        'value': '',
                    }
                    form_fields.append(form_field)
            
            result['forms'] = form_fields
            
        except Exception as e:
            self.logger.warning(f"检测表单字段失败: {e}")
    
    def _detect_form_patterns(self, text: str) -> List[Dict[str, Any]]:
        """检测文本中的表单模式"""
        patterns = []
        
        import re
        
        # 常见表单字段模式
        field_patterns = {
            'name': [r'姓名', r'Name', r'名称'],
            'id_card': [r'身份证号', r'ID', r'证件号码'],
            'phone': [r'电话', r'手机', r'Phone'],
            'email': [r'邮箱', r'Email', r'电子邮件'],
            'address': [r'地址', r'Address', r'住址'],
            'date': [r'日期', r'Date', r'时间'],
            'amount': [r'金额', r'Amount', r'数量'],
        }
        
        for field_type, type_patterns in field_patterns.items():
            for pattern in type_patterns:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                for match in matches:
                    patterns.append({
                        'type': field_type,
                        'pattern': pattern,
                        'position': match.span(),
                        'matched_text': match.group(),
                    })
        
        return patterns
    
    def _get_excel_type(self, file_path: Path) -> DocumentType:
        """根据文件扩展名确定Excel类型"""
        suffix = file_path.suffix.lower()
        if suffix == '.xlsx':
            return DocumentType.XLSX
        else:
            return DocumentType.XLS
    
    def get_supported_extensions(self) -> List[str]:
        """获取支持的文件扩展名"""
        return ['.xls', '.xlsx']
    
    def get_status(self) -> ProcessingStatus:
        """获取处理状态"""
        return self.status 